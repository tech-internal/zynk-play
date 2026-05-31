"""Generate FanVerse player endpoint URLs Excel workbook."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parents[1] / "docs" / "FanVerse_Player_Endpoint_URLs.xlsx"

HEADERS = ["Category", "Screen / Resource", "Method", "URL / Pattern", "Notes"]

ROWS = [
    # Frontend routes
    ("Frontend Routes", "Live stream + reels", "GET (page)", "/streaming", "React route → StreamingPage"),
    ("Frontend Routes", "Unity games hub", "GET (page)", "/gameplay", "React route → GamePlayPage"),
    # Live stream HLS
    (
        "Live Stream Player (HLS)",
        "IPL 2026 live stream manifest",
        "GET",
        "https://d1clrt8nxj7onv.cloudfront.net/live/myStream/playlist.m3u8",
        "Hardcoded in frontend/src/config/afgCricket.ts; used by StreamingPage",
    ),
    # Game players
    (
        "Game Player URLs (Unity WebGL)",
        "Afghan Soccer",
        "GET",
        "https://dd0pkk35jdxki.cloudfront.net/index.html",
        "Append ?userid={userId} when logged in; external launch (iframe blocked)",
    ),
    (
        "Game Player URLs (Unity WebGL)",
        "MI India Cricket",
        "GET",
        "https://d3kqwmfqrqx099.cloudfront.net/index.html",
        "Append ?userid={userId} when logged in",
    ),
    # Backend streaming API
    ("Backend API — Streaming", "List streams", "GET", "{API_BASE}/api/v1/streams", "AllowAny"),
    (
        "Backend API — Streaming",
        "Stream details",
        "GET",
        "{API_BASE}/api/v1/streams/{stream_id}",
        "AllowAny",
    ),
    (
        "Backend API — Streaming",
        "Request signed playback URL",
        "POST",
        "{API_BASE}/api/v1/streams/{stream_id}/access",
        "Auth required; returns signed_url, expires_in_seconds, session_id",
    ),
    (
        "Backend API — Streaming",
        "Start free trial",
        "POST",
        "{API_BASE}/api/v1/trial/start",
        "Auth required; body: content_id; returns signed_url (5 min)",
    ),
    (
        "Backend API — Streaming",
        "Signed playback URL (returned by access/trial)",
        "GET",
        "{STREAM_BASE_URL}/stream/{signed_url_token}",
        "Default STREAM_BASE_URL: https://stream.yourdomain.com",
    ),
    # Backend games API
    (
        "Backend API — Games",
        "List games",
        "GET",
        "{API_BASE}/api/v1/games",
        "Auth + game entitlement required",
    ),
    (
        "Backend API — Games",
        "Game details",
        "GET",
        "{API_BASE}/api/v1/games/{game_id}",
        "Auth + game entitlement required",
    ),
    (
        "Backend API — Games",
        "Launch game",
        "POST",
        "{API_BASE}/api/v1/games/{game_id}/launch",
        "Returns session_token and game_source from DB",
    ),
    # Reels
    (
        "Reels Video Player",
        "Local static file (preferred)",
        "GET",
        "/reels/{folderId}/{encodedFilePath}",
        "Served from frontend/public/reels/",
    ),
    (
        "Reels Video Player",
        "S3 direct fallback",
        "GET",
        "https://{bucket}.s3.{region}.amazonaws.com/{s3Key}",
        "Used when local file missing and bucket allows public read",
    ),
    (
        "Reels Video Player",
        "Backend proxy fallback",
        "GET",
        "{API_BASE}/api/v1/reels/stream?folder={folderId}&file={filePath}",
        "Example: .../api/v1/reels/stream?folder=folder2&file=Reels%2Fexample.mp4",
    ),
    # Env vars
    (
        "Environment Variables",
        "REACT_APP_API_URL",
        "—",
        "http://127.0.0.1:8000",
        "Frontend API base (default if unset)",
    ),
    (
        "Environment Variables",
        "SITE_URL",
        "—",
        "http://localhost:8000",
        "Django site base (.env.example default)",
    ),
    (
        "Environment Variables",
        "STREAM_BASE_URL",
        "—",
        "https://stream.yourdomain.com",
        "Signed live-stream playback base (.env.example default)",
    ),
]


def style_header(ws, row: int = 1) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws) -> None:
    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val:
                max_len = max(max_len, min(len(str(val)), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Player Endpoints"

    ws.append(HEADERS)
    style_header(ws)

    for row in ROWS:
        ws.append(list(row))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    autosize_columns(ws)

    # Summary sheet
    summary = wb.create_sheet("Quick Reference")
    summary.append(["Placeholder", "Default Value", "Description"])
    style_header(summary)
    summary.append(["{API_BASE}", "http://127.0.0.1:8000", "REACT_APP_API_URL or Django local"])
    summary.append(["{STREAM_BASE_URL}", "https://stream.yourdomain.com", "From .env STREAM_BASE_URL"])
    summary.append(["{stream_id}", "UUID", "From GET /api/v1/streams"])
    summary.append(["{game_id}", "UUID", "From GET /api/v1/games"])
    summary.append(["{signed_url_token}", "base64 token", "Returned by access/trial endpoints"])
    for row in summary.iter_rows(min_row=2, max_row=summary.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize_columns(summary)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
